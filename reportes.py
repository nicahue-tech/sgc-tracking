import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from bitacora import obtener_bitacora
from calculo import calcular_resultados
from correo_compensatorio import RUTA_SALIDA as RUTA_ARCHIVO_COMPENSATORIO

log = obtener_bitacora("reportes")

RUTA_PROYECTO = os.path.dirname(os.path.abspath(__file__))
RUTA_SALIDA = os.path.join(RUTA_PROYECTO, "reportes_generados")

FUENTE_ENLACE = Font(color="0563C1", underline="single")

RELLENO_ACUMULADO = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
RELLENO_PROYECTADO = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
RELLENO_MONTO_DIARIO = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

RELLENO_ROJO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RELLENO_NARANJO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RELLENO_VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def obtener_relleno_porcentaje_proyectado(porcentaje):
    if porcentaje < 70:
        return RELLENO_ROJO
    elif porcentaje < 100:
        return RELLENO_NARANJO
    else:
        return RELLENO_VERDE

ETIQUETAS_INDICADORES = [
    ("Monto acumulado", "acumulado"),
    ("Cantidad de operaciones", "cantidad_operaciones"),
    ("Cuota promedio", "cuota_promedio"),
    ("Promedio diario real", "promedio_diario_real"),
    ("Proyectado a cierre de mes", "proyectado_cierre"),
    ("Meta asignada", "meta"),
    ("Porcentaje de avance sobre la meta", "porcentaje_avance_meta"),
    ("Monto diario requerido para meta", "monto_diario_requerido"),
]


def escribir_enlaces_navegacion(hoja, nombres_hojas):
    columna = 1
    for nombre_hoja in nombres_hojas:
        celda = hoja.cell(row=1, column=columna, value=f"Ir a {nombre_hoja}")
        celda.hyperlink = f"#'{nombre_hoja}'!A1"
        celda.font = FUENTE_ENLACE
        columna += 1


def escribir_tabla_indicadores(hoja, fila_inicio, titulo, indicadores):
    hoja.cell(row=fila_inicio, column=1, value=titulo).font = Font(bold=True)
    fila = fila_inicio + 1

    hoja.cell(row=fila, column=1, value="Indicador").font = Font(bold=True)
    hoja.cell(row=fila, column=2, value="Valor").font = Font(bold=True)
    fila += 1

    RELLENOS_POR_CLAVE = {
        "acumulado": RELLENO_ACUMULADO,
        "proyectado_cierre": RELLENO_PROYECTADO,
        "monto_diario_requerido": RELLENO_MONTO_DIARIO
    }

    for etiqueta, clave in ETIQUETAS_INDICADORES:
        valor = indicadores.get(clave)
        if valor is None:
            valor_mostrado = "No aplica"
        elif clave == "cantidad_operaciones":
            valor_mostrado = valor
        elif clave == "porcentaje_avance_meta":
            valor_mostrado = f"{round(valor, 1)}%"
        else:
            valor_mostrado = round(valor)
        celda_etiqueta = hoja.cell(row=fila, column=1, value=etiqueta)
        celda_valor = hoja.cell(row=fila, column=2, value=valor_mostrado)
        if clave in RELLENOS_POR_CLAVE:
            celda_etiqueta.fill = RELLENOS_POR_CLAVE[clave]
            celda_valor.fill = RELLENOS_POR_CLAVE[clave]
        fila += 1

    meta = indicadores.get("meta")
    proyectado_cierre = indicadores.get("proyectado_cierre")
    if meta is not None and meta != 0:
        porcentaje_proyectado = (proyectado_cierre / meta) * 100
        valor_proyectado_mostrado = f"{round(porcentaje_proyectado, 1)}%"
        relleno_proyectado_porcentaje = obtener_relleno_porcentaje_proyectado(porcentaje_proyectado)
    else:
        valor_proyectado_mostrado = "No aplica"
        relleno_proyectado_porcentaje = None

    celda_etiqueta_proyectado = hoja.cell(row=fila, column=1, value="Proyectado sobre la meta (%)")
    celda_valor_proyectado = hoja.cell(row=fila, column=2, value=valor_proyectado_mostrado)
    if relleno_proyectado_porcentaje is not None:
        celda_etiqueta_proyectado.fill = relleno_proyectado_porcentaje
        celda_valor_proyectado.fill = relleno_proyectado_porcentaje
    fila += 1

    return fila + 1


def escribir_tabla_avance_diario(hoja, fila_inicio, titulo, avance_diario, clave_cantidad, clave_monto, clave_cuota):
    hoja.cell(row=fila_inicio, column=1, value=titulo).font = Font(bold=True)
    fila = fila_inicio + 1

    hoja.cell(row=fila, column=1, value="Fecha").font = Font(bold=True)
    hoja.cell(row=fila, column=2, value="Cantidad de operaciones (Q)").font = Font(bold=True)
    hoja.cell(row=fila, column=3, value="Monto del dia").font = Font(bold=True)
    hoja.cell(row=fila, column=4, value="Cuota promedio del dia").font = Font(bold=True)
    fila += 1

    total_cantidad = 0
    total_monto = 0

    for dia in avance_diario:
        cantidad_dia = dia[clave_cantidad]
        monto_dia = dia[clave_monto]
        cuota_dia = dia[clave_cuota]

        hoja.cell(row=fila, column=1, value=dia["fecha"].strftime("%d-%m-%Y"))
        hoja.cell(row=fila, column=2, value=cantidad_dia)
        hoja.cell(row=fila, column=3, value=round(monto_dia))
        hoja.cell(row=fila, column=4, value=round(cuota_dia))
        fila += 1

        total_cantidad += cantidad_dia
        total_monto += monto_dia

    cuota_promedio_total = total_monto / total_cantidad if total_cantidad > 0 else 0

    hoja.cell(row=fila, column=1, value="Total del mes").font = Font(bold=True)
    hoja.cell(row=fila, column=2, value=total_cantidad).font = Font(bold=True)
    hoja.cell(row=fila, column=3, value=round(total_monto)).font = Font(bold=True)
    hoja.cell(row=fila, column=4, value=round(cuota_promedio_total)).font = Font(bold=True)
    fila += 1

    return fila + 1


def escribir_tabla_ejecutivos(hoja, fila_inicio, titulo, ejecutivos, tipo_indicador):
    hoja.cell(row=fila_inicio, column=1, value=titulo).font = Font(bold=True)
    fila = fila_inicio + 1

    encabezados = [
        "Ejecutivo", "Supervisor", "Tipo", "Monto acumulado", "Cantidad (Q)",
        "Cuota promedio", "Promedio diario real", "Proyectado a cierre",
        "Meta", "Monto diario requerido", "Porcentaje avance", "Proyectado %"
    ]
    COLUMNA_ACUMULADO = 4
    COLUMNA_PROYECTADO = 8
    COLUMNA_MONTO_DIARIO = 10
    COLUMNA_PROYECTADO_PORCENTAJE = 12
    RELLENOS_POR_COLUMNA = {
        COLUMNA_ACUMULADO: RELLENO_ACUMULADO,
        COLUMNA_PROYECTADO: RELLENO_PROYECTADO,
        COLUMNA_MONTO_DIARIO: RELLENO_MONTO_DIARIO
    }
    for columna, encabezado in enumerate(encabezados, start=1):
        celda_encabezado = hoja.cell(row=fila, column=columna, value=encabezado)
        celda_encabezado.font = Font(bold=True)
        if columna in RELLENOS_POR_COLUMNA:
            celda_encabezado.fill = RELLENOS_POR_COLUMNA[columna]
    fila += 1

    lista_ordenada = sorted(
        ejecutivos.items(),
        key=lambda item: item[1][tipo_indicador]["acumulado"],
        reverse=True
    )

    total_acumulado = 0
    total_cantidad = 0
    total_promedio_diario = 0
    total_proyectado = 0
    total_meta = 0
    total_monto_diario_requerido = 0
    hay_meta = False

    for nombre_ejecutivo, datos_ejecutivo in lista_ordenada:
        indicadores = datos_ejecutivo[tipo_indicador]
        meta_mostrada = "No aplica" if indicadores["meta"] is None else round(indicadores["meta"])
        monto_diario_mostrado = "No aplica" if indicadores["monto_diario_requerido"] is None else round(indicadores["monto_diario_requerido"])
        porcentaje_mostrado = "No aplica" if indicadores["porcentaje_avance_meta"] is None else f"{round(indicadores['porcentaje_avance_meta'], 1)}%"

        if indicadores["meta"] is not None and indicadores["meta"] != 0:
            porcentaje_proyectado = (indicadores["proyectado_cierre"] / indicadores["meta"]) * 100
            porcentaje_proyectado_mostrado = f"{round(porcentaje_proyectado, 1)}%"
            relleno_proyectado_porcentaje = obtener_relleno_porcentaje_proyectado(porcentaje_proyectado)
        else:
            porcentaje_proyectado_mostrado = "No aplica"
            relleno_proyectado_porcentaje = None

        valores_fila = [
            nombre_ejecutivo,
            datos_ejecutivo["supervisor"],
            datos_ejecutivo["tipo"],
            round(indicadores["acumulado"]),
            indicadores["cantidad_operaciones"],
            round(indicadores["cuota_promedio"]),
            round(indicadores["promedio_diario_real"]),
            round(indicadores["proyectado_cierre"]),
            meta_mostrada,
            monto_diario_mostrado,
            porcentaje_mostrado,
            porcentaje_proyectado_mostrado
        ]
        for columna, valor in enumerate(valores_fila, start=1):
            celda = hoja.cell(row=fila, column=columna, value=valor)
            if columna in RELLENOS_POR_COLUMNA:
                celda.fill = RELLENOS_POR_COLUMNA[columna]
            if columna == COLUMNA_PROYECTADO_PORCENTAJE and relleno_proyectado_porcentaje is not None:
                celda.fill = relleno_proyectado_porcentaje
        fila += 1

        total_acumulado += indicadores["acumulado"]
        total_cantidad += indicadores["cantidad_operaciones"]
        total_promedio_diario += indicadores["promedio_diario_real"]
        total_proyectado += indicadores["proyectado_cierre"]
        if indicadores["meta"] is not None:
            hay_meta = True
            total_meta += indicadores["meta"]
            total_monto_diario_requerido += indicadores["monto_diario_requerido"]

    cuota_promedio_total = total_acumulado / total_cantidad if total_cantidad > 0 else 0
    meta_total_mostrada = round(total_meta) if hay_meta else "No aplica"
    monto_diario_total_mostrado = round(total_monto_diario_requerido) if hay_meta else "No aplica"
    porcentaje_total_mostrado = f"{round((total_acumulado / total_meta) * 100, 1)}%" if hay_meta and total_meta != 0 else "No aplica"

    if hay_meta and total_meta != 0:
        porcentaje_proyectado_total = (total_proyectado / total_meta) * 100
        porcentaje_proyectado_total_mostrado = f"{round(porcentaje_proyectado_total, 1)}%"
        relleno_proyectado_total = obtener_relleno_porcentaje_proyectado(porcentaje_proyectado_total)
    else:
        porcentaje_proyectado_total_mostrado = "No aplica"
        relleno_proyectado_total = None

    valores_totales = [
        "Total", "", "",
        round(total_acumulado), total_cantidad, round(cuota_promedio_total),
        round(total_promedio_diario), round(total_proyectado),
        meta_total_mostrada, monto_diario_total_mostrado, porcentaje_total_mostrado,
        porcentaje_proyectado_total_mostrado
    ]
    for columna, valor in enumerate(valores_totales, start=1):
        celda = hoja.cell(row=fila, column=columna, value=valor)
        celda.font = Font(bold=True)
        if columna in RELLENOS_POR_COLUMNA:
            celda.fill = RELLENOS_POR_COLUMNA[columna]
        if columna == COLUMNA_PROYECTADO_PORCENTAJE and relleno_proyectado_total is not None:
            celda.fill = relleno_proyectado_total
    fila += 1

    return fila + 1


def generar_hoja_global(libro, resultados):
    hoja = libro.active
    hoja.title = "Global"

    hoja.column_dimensions["A"].width = 32
    hoja.column_dimensions["B"].width = 26
    hoja.column_dimensions["C"].width = 20
    hoja.column_dimensions["D"].width = 24

    hoja.cell(row=3, column=1, value="GLOBAL").font = Font(bold=True, size=13)
    fila = 4
    hoja.cell(row=fila, column=1, value=f"Fecha y hora de corte: {resultados['fecha_corte'].strftime('%d-%m-%Y %H:%M')}").font = Font(bold=True)
    fila += 1
    hoja.cell(row=fila, column=1, value=f"Dias habiles transcurridos: {resultados['dias_habiles_transcurridos']} de {resultados['dias_habiles_totales_mes']}")
    fila += 2

    fila = escribir_tabla_indicadores(hoja, fila, "Acuerdos", resultados["global"]["acuerdo"])
    fila = escribir_tabla_indicadores(hoja, fila, "Reajustes", resultados["global"]["reajuste"])

    fila = escribir_tabla_avance_diario(
        hoja, fila, "Avance diario de Acuerdos", resultados["avance_diario"],
        "cantidad_acuerdo", "monto_acuerdo", "cuota_promedio_acuerdo"
    )
    escribir_tabla_avance_diario(
        hoja, fila, "Avance diario de Reajustes", resultados["avance_diario"],
        "cantidad_reajuste", "monto_reajuste", "cuota_promedio_reajuste"
    )

    return hoja


def generar_hojas_supervisores(libro, resultados):
    hojas_creadas = []
    for supervisor, datos_supervisor in resultados["supervisores"].items():
        hoja = libro.create_sheet(title=supervisor[:31])
        hoja.column_dimensions["A"].width = 32
        hoja.column_dimensions["B"].width = 22
        hoja.column_dimensions["C"].width = 12
        hoja.column_dimensions["D"].width = 16
        hoja.column_dimensions["E"].width = 12
        hoja.column_dimensions["F"].width = 16
        hoja.column_dimensions["G"].width = 18
        hoja.column_dimensions["H"].width = 18
        hoja.column_dimensions["I"].width = 14
        hoja.column_dimensions["J"].width = 16
        hoja.column_dimensions["K"].width = 20
        hoja.column_dimensions["L"].width = 16

        hoja.cell(row=3, column=1, value=f"Supervisor: {supervisor}").font = Font(bold=True, size=13)
        fila = 5

        fila = escribir_tabla_indicadores(hoja, fila, "Acuerdos del equipo (total)", datos_supervisor["acuerdo"])
        fila = escribir_tabla_indicadores(hoja, fila, "Reajustes del equipo (total)", datos_supervisor["reajuste"])

        ejecutivos_del_equipo = {
            nombre: datos for nombre, datos in resultados["ejecutivos"].items()
            if datos["supervisor"] == supervisor
        }

        fila = escribir_tabla_ejecutivos(hoja, fila, "Acuerdos individualizados del equipo", ejecutivos_del_equipo, "acuerdo")
        escribir_tabla_ejecutivos(hoja, fila, "Reajustes individualizados del equipo", ejecutivos_del_equipo, "reajuste")

        log.info(f"Hoja de supervisor generada: {supervisor}")
        hojas_creadas.append(hoja)

    return hojas_creadas


def generar_hoja_ejecutivos(libro, resultados):
    hoja = libro.create_sheet(title="Ejecutivos")
    anchos = [32, 22, 12, 16, 12, 16, 18, 18, 14, 16, 20, 16]
    letras_columnas = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
    for letra, ancho in zip(letras_columnas, anchos):
        hoja.column_dimensions[letra].width = ancho

    hoja.cell(row=3, column=1, value="EJECUTIVOS").font = Font(bold=True, size=13)
    fila = escribir_tabla_ejecutivos(hoja, 5, "Acuerdos por ejecutivo", resultados["ejecutivos"], "acuerdo")
    escribir_tabla_ejecutivos(hoja, fila, "Reajustes por ejecutivo", resultados["ejecutivos"], "reajuste")
    log.info("Hoja de ejecutivos generada correctamente")
    return hoja


def generar_hoja_cumplimiento_ejecutivos(libro):
    hoja = libro.create_sheet(title="Cumplimiento Ejecutivos")
    anchos = [14, 34, 24, 18, 20, 14]
    letras_columnas = ["A", "B", "C", "D", "E", "F"]
    for letra, ancho in zip(letras_columnas, anchos):
        hoja.column_dimensions[letra].width = ancho

    hoja.cell(row=3, column=1, value="CUMPLIMIENTO EJECUTIVOS").font = Font(bold=True, size=13)
    hoja.cell(
        row=4, column=1,
        value=(
            "Hoja informativa, tomada tal cual del reporte diario de compensatorio "
            "que llega por correo. No se cruza ni se relaciona con los nombres "
            "normalizados del resto del sistema."
        )
    )

    fila = 6

    if not os.path.isfile(RUTA_ARCHIVO_COMPENSATORIO):
        hoja.cell(
            row=fila, column=1,
            value="Sin datos disponibles todavia, no se ha descargado el reporte de hoy."
        )
        log.info("Hoja Cumplimiento Ejecutivos generada sin datos, no se encontro el archivo temporal.")
        return hoja

    libro_origen = load_workbook(RUTA_ARCHIVO_COMPENSATORIO, data_only=True)
    hoja_origen = libro_origen.active

    encabezados = [
        "Supervisor", "Ejecutivo", "% Cumplimiento Productivo",
        "% Compensacion", "% Cumplimiento Total", "% Proyectado"
    ]
    for columna, encabezado in enumerate(encabezados, start=1):
        hoja.cell(row=fila, column=columna, value=encabezado).font = Font(bold=True)
    fila += 1

    filas_origen = [
        fila_datos for fila_datos in hoja_origen.iter_rows(min_row=2, values_only=True)
        if fila_datos and fila_datos[0] is not None
    ]

    for fila_datos in filas_origen:
        supervisor, ejecutivo, cumpl_productivo, compensacion, cumpl_total, proyectado = fila_datos[:6]
        hoja.cell(row=fila, column=1, value=supervisor)
        hoja.cell(row=fila, column=2, value=ejecutivo)
        for columna, valor in zip([3, 4, 5, 6], [cumpl_productivo, compensacion, cumpl_total, proyectado]):
            celda = hoja.cell(row=fila, column=columna, value=valor)
            celda.number_format = "0.0%"
        fila += 1

    log.info(f"Hoja Cumplimiento Ejecutivos generada con {len(filas_origen)} ejecutivos.")
    return hoja


def generar_reporte_completo(resultados):
    libro = Workbook()

    hoja_global = generar_hoja_global(libro, resultados)
    hojas_supervisores = generar_hojas_supervisores(libro, resultados)
    hoja_ejecutivos = generar_hoja_ejecutivos(libro, resultados)
    hoja_cumplimiento_ejecutivos = generar_hoja_cumplimiento_ejecutivos(libro)

    todas_las_hojas = [hoja_global] + hojas_supervisores + [hoja_ejecutivos, hoja_cumplimiento_ejecutivos]
    nombres_hojas = [hoja.title for hoja in todas_las_hojas]

    for hoja in todas_las_hojas:
        escribir_enlaces_navegacion(hoja, nombres_hojas)

    os.makedirs(RUTA_SALIDA, exist_ok=True)
    ruta_archivo = os.path.join(RUTA_SALIDA, "dashboard_sgc_tracking.xlsx")
    libro.save(ruta_archivo)
    log.info(f"Reporte completo generado correctamente en: {ruta_archivo}")
    return ruta_archivo


if __name__ == "__main__":
    resultados = calcular_resultados()
    ruta = generar_reporte_completo(resultados)
    log.info(f"Prueba de reportes.py finalizada, archivo en {ruta}")
